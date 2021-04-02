import spellchecker as sc
from openpyxl import Workbook, load_workbook
    
if __name__ == "__main__":
    print('Removing duplicate text rows from text.xlsx...')
    wb = load_workbook('text.xlsx')
    ws = wb.worksheets[0]
    sentences = []
    # Create list of non-duplicate sentences
    sc = sc.SpellChecker(_spacy=False)
    col = sc._get_column_with_title(ws, 'Text', col_letter=True)
    col_int = sc._get_column_with_title(ws, 'Text')
    if col:
        row_start_idx = 1
    else:
        col = 'A'
        col_int = 1
        row_start_idx = 0
    
    for cell in ws[col][row_start_idx:]:
        if cell.value not in sentences:
            sentences.append(cell.value)
        cell.value = None
    # Populate column with sentences
    for i, s in enumerate(sentences, start=1):
        ws.cell(row_start_idx+i, col_int).value = s
    
    print(f'Removed {len(ws[col][row_start_idx:]) - len(sentences)} duplicate lines!')
    wb.save('text.xlsx')
    