import difflib
import re
import sys
import time
import urllib
from collections import defaultdict
from datetime import timedelta

import contextualSpellCheck
import enchant
import requests
import spacy
from bs4 import BeautifulSoup
from english_words import (english_words_lower_alpha_set,
                           english_words_lower_set)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from tqdm import tqdm

headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
           '(KHTML, like Gecko) Chrome/88.0.4324.190 Safari/537.36'}


class SpellChecker:
    def __init__(self, _spacy=True):
        # Input file must contain text to spellcheck
        # on the first sheet and listed in column A,
        # or with the column header 'Text'
        self.input_file = 'text.xlsx'
        # Output file shows all corrected and non-corrected
        # words with context
        self.result_file = 'result.xlsx'
        # Output text file with all spelling corrections applied.
        # This file is created with the 'apply_user_actions' method/script
        self.text_output_file = 'text-output.xlsx'
        # Dictionary file must contain a custom word list
        # on the first sheet and listed in column A.
        # Words can be seperated in the same cell by pipes ' | '
        self.word_dict_file = 'dictionary.xlsx'

        self.word_dict = {}
        self.enchant_dict_US = enchant.Dict("en_US")
        self.enchant_dict_GB = enchant.Dict("en_GB")

        # NLP Contextual Spell Checker
        if _spacy:
            tqdm.write('Loading Spacy pipeline...')
            self.nlp = spacy.load("en_core_web_sm")
            self.nlp.add_pipe('contextual spellchecker')
            tqdm.write('Spacy loaded.')

        self.corrected_words = []
        self.not_corrected_words = []

        self.count = defaultdict(int)

        self.session = requests.Session()

    def spell_check_text(self, num_context_words=5, auto=True, suggest=True, debug=False, google_sc=True):
        '''
        Checks input sentences aginst custom word list as well as other english words lists
        and outputs result to the output file
        '''
        if debug:
            self.words = []

        if not self.word_dict:
            self._load_word_dict()

        workbook = load_workbook(self.input_file)
        worksheet = workbook.worksheets[0]

        # Find column with 'text' as title cell
        tqdm.write('Finding "Text" column...')
        
        col = self._get_column_with_title(worksheet, 'text', col_letter=True)
        if not col:
        # if not (col := self._get_column_with_title(worksheet, 'text', col_letter=True)):
            col = get_column_letter(1)
            row_start_idx = 0  # Row index to start from
            tqdm.write(f'No column with title "text", defaulting to column {col}')
        else:
            row_start_idx = 1  # Ignore title row if it exists
            tqdm.write(f'Column {col} has title "text"')

        tqdm.write('Spell checking text...')
        start_time = time.time()
        # Iterate through each cell in text column
        for cell in tqdm(worksheet[col][row_start_idx:]):
            if not cell.value:
                continue

            text = str(cell.value)

            words = text.split()
            for i, w in enumerate(words):
                try:
                    # Remove punctuation
                    if w[-1] in [',', '.', ';', ':', '/', '-', '!', '?', '%']:
                        w = w[:-1]
                    # Remove brackets
                    if w[0] in ['(', '[', '{']:
                        w = w[1:]
                    if w[-1] in [')', ']', '}']:
                        w = w[:-1]

                except IndexError:
                    continue

                # Track words checked
                self.count['words-checked'] += 1
                if debug:
                    self.words.append(w)

                # Ignore words with numbers
                if re.search('[0-9]+', w):
                    continue
                # Ignore short, all caps words (Likely tickers)
                if len(w) <= 5 and (w == w.upper()):
                    continue
                # Ignore capiltalized words (Names)
                if w == w.capitalize():
                    continue

                if not w:
                    continue

                word = w.lower()

                # Check dictionaries
                if word in self.word_dict:
                    continue
                if (word in english_words_lower_alpha_set or
                        word in english_words_lower_set):
                    continue
                if self.enchant_dict_US.check(word):
                    continue
                if self.enchant_dict_GB.check(word):
                    continue

                self.count['words-misspelled'] += 1

                # Get sentence context for word
                first_context_index = max(0, i-num_context_words)
                last_context_index = min(i+num_context_words+1, len(words))
                context_words = [
                    w for w in words[first_context_index: last_context_index]]

                if first_context_index > 0:
                    context_words.insert(0, '...')
                if last_context_index < len(words):
                    context_words.append('...')

                context = " ".join(context_words)

                google_search_context_url = 'http://www.google.com/search?q=' + \
                    urllib.parse.quote_plus(context)
                google_search_word_url = 'http://www.google.com/search?q=' + \
                    urllib.parse.quote_plus(w)

                # Get word suggestions from enchant for misspelt word
                suggestions = []
                if suggest:
                    suggestions = self.enchant_dict_US.suggest(word)
                    # Use up to 3 suggestions
                    suggestions = suggestions[:min(3, len(suggestions))]

                if auto:
                    # Context spell check for suggested words
                    for suggestion in suggestions:
                        new_text = text.replace(w, suggestion)
                        res = self._spacy_spellcheck(suggestion, new_text)
                        # If suggested word is not flagged as misspelt
                        if res:
                            # Print correction to terminal
                            _f_new_context = self._get_f_new_context(
                                context, w, suggestion)
                            tqdm.write(_f_new_context)
                            # Store correction data for outputting
                            self.corrected_words.append({'word': w,
                                                         'correction': suggestion,
                                                         'row': cell.row,
                                                         'context': context,
                                                         'suggestions': suggestions,
                                                         'search_context_url': google_search_context_url,
                                                         'search_word_url': google_search_word_url})
                            self.count['words-corrected'] += 1
                            break
                    else:
                        if google_sc and suggest:
                            # Get google searches suggested word correction
                            _google_output = self._get_google_correction(google_search_context_url, word=w)
                            if _google_output:
                                google_suggested_words = _google_output
                            else:
                                google_suggested_words = []
                            # google_suggested_words = _ if (
                            #     _ := self._get_google_correction(google_search_context_url, word=w)) else []

                            if len(google_suggested_words) > 1:
                                try:
                                    google_suggested_word = difflib.get_close_matches(
                                        w, google_suggested_words, cutoff=0.4)[0]
                                except IndexError:
                                    continue
                            elif len(google_suggested_words) == 1:
                                google_suggested_word = google_suggested_words[0]

                            else:
                                self.not_corrected_words.append({'word': w,
                                                                 'row': cell.row,
                                                                 'context': context,
                                                                 'suggestions': suggestions,
                                                                 'search_context_url': google_search_context_url,
                                                                 'search_word_url': google_search_word_url})
                                self.count['google-words-not-corrected'] += 1
                                continue

                            _f_new_context = self._get_f_new_context(
                                context, w, google_suggested_word)
                            tqdm.write('Google Correction:' + _f_new_context)
                            self.corrected_words.append({'word': w,
                                                         'correction': google_suggested_word,
                                                         'row': cell.row,
                                                         'context': context,
                                                         'suggestions': suggestions,
                                                         'search_context_url': google_search_context_url,
                                                         'search_word_url': google_search_word_url})
                            self.count['google-words-corrected'] += 1
                            continue

                        self.not_corrected_words.append({'word': w,
                                                         'row': cell.row,
                                                         'context': context,
                                                         'suggestions': suggestions,
                                                         'search_context_url': google_search_context_url,
                                                         'search_word_url': google_search_word_url})
                        self.count['words-not-corrected'] += 1
                else:
                    self.not_corrected_words.append({'word': w,
                                                     'row': cell.row,
                                                     'context': context,
                                                     'suggestions': suggestions,
                                                     'search_context_url': google_search_context_url,
                                                     'search_word_url': google_search_word_url})

        secs_taken = time.time() - start_time
        f_time = format_time(secs=secs_taken)
        print(f'\nText spellchecked in {f_time}\n')
        print(f'Checked {self.count["words-checked"]} words.')
        print(f'Found {self.count["words-misspelled"]} misspelt words.')
        print(f'Corrected {self.count["words-corrected"]} words.')
        print(f'Unable to correct {self.count["words-not-corrected"]} words.')
        print(
            f'Google corrected {self.count["google-words-corrected"]} words.')
        print(
            f'Google could not correct {self.count["google-words-not-corrected"]} words.')

        print('\nSaving results file...')
        self._output_spacy()
        if debug:
            self._output_debug()
        print('Results saved!')

    def _spacy_spellcheck(self, word, text):
        ''' Returns True if input word passes Spacy spellcheck, else False '''
        doc = self.nlp(text)
        # check if corrected word is still marked as mistake
        if doc._.performed_spellCheck:
            misspelt_spacy = [w.text for w,
                              s in doc._.score_spellCheck.items()]
            if word in misspelt_spacy:
                return False
        return True

    def _get_column_with_title(self, worksheet, text, exact_match=False, col_letter=False):
        # Check cells in row 1 for matching text
        for j in range(1, worksheet.max_column+1):
            cell = worksheet.cell(1, j)
            cell_text = str(cell.value)

            if exact_match:
                if cell_text == text:
                    return get_column_letter(j) if col_letter else j
            else:
                if cell_text.strip().lower() == text.strip().lower():
                    return get_column_letter(j) if col_letter else j

        return False

    def _get_google_correction(self, url, word=None):
        ''' Returns the corrected word from google search "Did you mean: ..." '''
        tqdm.write(f'Googling word "{word}"...')
        time.sleep(3)  # Prevent spam/I.P. block
        r = self.session.get(url, headers=headers)
        soup = BeautifulSoup(r.text, 'html.parser')

        if 300 >= r.status_code < 200:
            tqdm.write(
                f'WARN: Request returned status code: {r.status_code}. Your IP address may be blocked by google.')
            return

        main_content = soup.find('div', id='main')
        if not main_content:
            tqdm.write(
                f'WARN: No content found on google search. You may have hit a spam block/captcha.')
            return

        for a_tag in main_content.find_all('a', href=re.compile('/search')):
            if ('Did you mean' in a_tag.parent.text or
                    'Showing results for' in a_tag.parent.text):
                suggested_words = [b.text for b in a_tag.find_all('b')]
                return suggested_words
        return

    def _get_f_new_context(self, context, old_word, new_word):
        c_words = []
        for c_word in context.split():
            word_match = re.match(r'[^0-9a-zA-Z]*(.+?)[^0-9a-zA-Z]*$', c_word)
            if word_match.group(1).lower() == old_word.lower():
                c_words.append(c_word.replace(word_match.group(
                    1), f'{strikethrough(old_word)} \033[1m{new_word}\033[0m'))
                continue
            c_words.append(c_word)
        return ' '.join(c_words)

    def _load_word_dict(self):
        '''
        Load custom word dictionary from file

        '''
        workbook = load_workbook(self.word_dict_file)
        worksheet = workbook.worksheets[0]

        for cell in worksheet['A']:
            # Split words/phrases in the same cell which are seperated by pipes '|'
            cell_value = str(cell.value)
            words = cell_value.split('|')

            for w in words:
                word = w.strip().lower()
                self.word_dict[word] = True

    def _output_spacy(self):
        workbook = Workbook()
        # Non-corrected words output
        not_corrected_ws = workbook.create_sheet('Not Corrected', 0)
        not_corrected_ws['A1'] = 'User Action'
        not_corrected_ws['A1'].font = Font(bold=True)
        not_corrected_ws.column_dimensions['A'].alignment = Alignment(
            horizontal='center')
        not_corrected_ws['B1'] = 'Word'
        not_corrected_ws['B1'].font = Font(bold=True)
        not_corrected_ws['C1'] = 'Row'
        not_corrected_ws['C1'].font = Font(bold=True)
        not_corrected_ws['D1'] = 'Original Context'
        not_corrected_ws['D1'].font = Font(bold=True)
        not_corrected_ws['E1'] = 'Suggestions'
        not_corrected_ws['E1'].font = Font(bold=True)
        # Set column widths
        not_corrected_ws.column_dimensions['A'].width = 20
        not_corrected_ws.column_dimensions['B'].width = 25
        not_corrected_ws.column_dimensions['C'].width = 5
        not_corrected_ws.column_dimensions['D'].width = 100
        not_corrected_ws.column_dimensions['E'].width = 20
        not_corrected_ws.column_dimensions['F'].width = 20
        not_corrected_ws.column_dimensions['G'].width = 20

        for i, result in enumerate(self.not_corrected_words, start=1):
            not_corrected_ws.cell(i+1, 2).value = result['word']
            not_corrected_ws.cell(i+1, 2).hyperlink = result['search_word_url']
            not_corrected_ws.cell(i+1, 2).style = 'Hyperlink'

            not_corrected_ws.cell(i+1, 3).value = result['row']

            not_corrected_ws.cell(i+1, 4).value = result['context']
            not_corrected_ws.cell(
                i+1, 4).hyperlink = result['search_context_url']
            not_corrected_ws.cell(i+1, 4).style = 'Hyperlink'
            try:
                not_corrected_ws.cell(i+1, 5).value = result['suggestions'][0]
                not_corrected_ws.cell(i+1, 6).value = result['suggestions'][1]
                not_corrected_ws.cell(i+1, 7).value = result['suggestions'][2]
            except IndexError:
                pass

        # Corrected words output
        corrected_ws = workbook.create_sheet('Corrected', 1)
        corrected_ws['A1'] = 'User Action'
        corrected_ws['A1'].font = Font(bold=True)
        corrected_ws.column_dimensions['A'].alignment = Alignment(
            horizontal='center')
        corrected_ws['B1'] = 'Word'
        corrected_ws['B1'].font = Font(bold=True)
        corrected_ws['C1'] = 'Correction'
        corrected_ws['C1'].font = Font(bold=True)
        corrected_ws['D1'] = 'Row'
        corrected_ws['D1'].font = Font(bold=True)
        corrected_ws['E1'] = 'Original Context'
        corrected_ws['E1'].font = Font(bold=True)
        corrected_ws['F1'] = 'Suggestions'
        corrected_ws['F1'].font = Font(bold=True)
        # Set column widths
        corrected_ws.column_dimensions['A'].width = 20
        corrected_ws.column_dimensions['B'].width = 25
        corrected_ws.column_dimensions['C'].width = 25
        corrected_ws.column_dimensions['D'].width = 5
        corrected_ws.column_dimensions['E'].width = 100
        corrected_ws.column_dimensions['F'].width = 20
        corrected_ws.column_dimensions['G'].width = 20
        corrected_ws.column_dimensions['H'].width = 20

        # start=2 to skip header row
        for i, result in enumerate(self.corrected_words, start=2):
            corrected_ws.cell(i, 2).value = result['word']
            corrected_ws.cell(i, 2).hyperlink = result['search_word_url']
            corrected_ws.cell(i, 2).style = 'Hyperlink'

            corrected_ws.cell(i, 3).value = result['correction']
            corrected_ws.cell(i, 4).value = result['row']

            corrected_ws.cell(i, 5).value = result['context']
            corrected_ws.cell(i, 5).hyperlink = result['search_context_url']
            corrected_ws.cell(i, 5).style = 'Hyperlink'
            try:
                corrected_ws.cell(i, 6).value = result['suggestions'][0]
                corrected_ws.cell(i, 7).value = result['suggestions'][1]
                corrected_ws.cell(i, 8).value = result['suggestions'][2]
            except IndexError:
                pass

        workbook.save(self.result_file)

    def _output_debug(self):
        wb = Workbook()
        words_ws = wb.create_sheet('Words Checked', 0)

        for i, word in enumerate(self.words, start=1):
            words_ws.cell(i, 1).value = word

        wb.save('debug.xlsx')

    def apply_user_actions(self):
        ''' Process user actions in result.xlsx file '''
        tqdm.write('Applying word corrections and custom user actions to text...')
        result_wb = load_workbook(self.result_file)
        not_corrected_ws = result_wb['Not Corrected']
        corrected_ws = result_wb['Corrected']

        words_to_add = set()  # Set of words to add to dictionary file
        words_to_change = []  # List of dicts containing old word, new word, sentence line number
        words_to_delete = []  # List of dicts containing word to delete, sentence line number
        # Iterate result rows
        for i in range(2, not_corrected_ws.max_row):  # Skip title row
            user_input = not_corrected_ws.cell(i, 1).value
            misspelled_word = not_corrected_ws.cell(i, 2).value
            text_row = not_corrected_ws.cell(i, 3).value
            # context = not_corrected_ws.cell(i, 4).value
            suggestions = []
            for j in range(5, not_corrected_ws.max_column+1):
                _s = not_corrected_ws.cell(i, j).value
                if _s:
                    suggestions.append(_s)
                # if (s := not_corrected_ws.cell(i, j).value):
                #     suggestions.append(s)
            # Ignore empty user actions on not corrected worksheet
            if not user_input:
                continue
            # Change word in text to user selected suggestion
            try:
                fixed_word = suggestions[int(user_input)]
                words_to_change.append(
                    {'line': text_row, 'old_word': misspelled_word, 'new_word': fixed_word})
                continue
            except ValueError:
                pass
            except IndexError:
                tqdm.write(
                    f'WARN: word suggestion index ({user_input}) out of suggestion bounds on "{self.result_file} line {i}"')

            # Add word to dictionary
            if (user_input.lower() == 'a' or
                    user_input.lower() == 'add'):
                if misspelled_word not in words_to_add:
                    words_to_add.add(misspelled_word)
                continue

            # Delete word
            if (user_input.lower() == 'd' or
                    user_input.lower() == 'del'):
                words_to_delete.append(
                    {'line': text_row, 'word': misspelled_word})
                continue

            # Change word
            words_to_change.append(
                {'line': text_row, 'old_word': misspelled_word, 'new_word': user_input})

        for i in range(2, corrected_ws.max_row):  # Skip title row
            user_input = corrected_ws.cell(i, 1).value
            misspelled_word = corrected_ws.cell(i, 2).value
            corrected_word = corrected_ws.cell(i, 3).value
            text_row = corrected_ws.cell(i, 4).value
            suggestions = []
            for j in range(6, corrected_ws.max_column+1):
                _s = corrected_ws.cell(i, j).value
                if _s:
                    suggestions.append(_s)
                # if (s := corrected_ws.cell(i, j).value):
                #     suggestions.append(s)
            # On corrected worksheet, if user action is empty then apply corrected word to text output
            if not user_input:
                words_to_change.append(
                    {'line': text_row, 'old_word': misspelled_word, 'new_word': corrected_word})
                continue

            # Change word in text to user selected suggestion
            try:
                fixed_word = suggestions[int(user_input)]
                words_to_change.append(
                    {'line': text_row, 'old_word': misspelled_word, 'new_word': fixed_word})
                continue
            except ValueError:
                pass
            except IndexError:
                tqdm.write(
                    f'WARN: word suggestion index ({user_input}) out of suggestion bounds on "{self.result_file} line {i}"')

            # Add word to dictionary
            if (user_input.lower() == 'a' or
                    user_input.lower() == 'add'):
                if misspelled_word not in words_to_add:
                    words_to_add.add(misspelled_word)
                continue

            # Delete word
            if (user_input.lower() == 'd' or
                    user_input.lower() == 'del'):
                words_to_delete.append(
                    {'line': text_row, 'word': misspelled_word})
                continue

            # Change word
            words_to_change.append(
                {'line': text_row, 'old_word': misspelled_word, 'new_word': user_input})

        # Apply word changes to text file and output to a new file
        text_output_wb = load_workbook(self.input_file)
        # Replace misspelled words with correct words in input text file
        text_output_ws = text_output_wb.worksheets[0]
        col = self._get_column_with_title(text_output_ws, 'text')
        col = col if col else 1

        # Change/correct words in text file
        for change in words_to_change:
            cell = text_output_ws.cell(change['line'], col)

            regex = fr'[^0-9a-zA-Z]+({change["old_word"]})[^0-9a-zA-Z]+'
            pattern = re.compile(regex)
            cell.value = ' ' + cell.value + ' '
            word_match = pattern.search(cell.value)
            if not word_match:
                continue

            to_replace = word_match.group(0)
            word_group = word_match.group(1)
            replace_with = to_replace.replace(word_group, change['new_word'])
            cell.value = str(cell.value).replace(
                to_replace, replace_with).strip()
            self.count['user-words-corrected'] += 1

        # Add words to dictionary
        dict_workbook = load_workbook(self.word_dict_file)
        dict_worksheet = dict_workbook.worksheets[0]
        for word in words_to_add:
            dict_worksheet.append([word])
            self.count['user-words-added-to-dictionary'] += 1
        dict_workbook.save(self.word_dict_file)

        # Delete words
        for word in words_to_delete:
            cell = text_output_ws.cell(word['line'], col)

            regex = fr'[^0-9a-zA-Z]+({word["word"]})[^0-9a-zA-Z]+'
            pattern = re.compile(regex)
            cell.value = ' ' + cell.value + ' '
            word_match = pattern.search(cell.value)
            if not word_match:
                continue

            to_replace = word_match.group(0)
            word_group = word_match.group(1)
            cell.value = str(cell.value).replace(to_replace, '').strip()
            self.count['user-words-deleted'] += 1

        # Save output text to file
        text_output_wb.save(self.text_output_file)

        print(
            f'\nUser applied {self.count["user-words-corrected"]} word corrections.')
        print(
            f'User added {self.count["user-words-added-to-dictionary"]} words to Dictionary.')
        print(f'\nUser deleted {self.count["user-words-deleted"]} words.')
        return


### HELPER FUNCTIONS ###

def print_process_time(msg, func, *args, **kwargs):
    ''' Prints time taken to run a function '''
    start_time = time.time()
    ret = func(*args, **kwargs)
    time_taken = time.time() - start_time
    print(f'{msg} in {time_taken:.2f}s')
    return ret


def strikethrough(text):
    s = ''
    for c in text:
        s += '\u0336' + c
    return s


def format_time(hours=0, mins=0, secs=0):
    ''' Format time into the format 7h5m32s / 17m12s / 45s '''
    m, s = divmod(secs, 60)
    m += mins
    h, m = divmod(m, 60)
    f_string = ''
    if h > 0:
        f_string += f'{h:.0f}h '
        f_string += f'{m:.0f}m '
    elif m > 0:
        f_string += f'{m:.0f}m '
    f_string += f'{s:.1f}s'
    return f_string

if __name__ == "__main__":
    
    sc_kwargs = {}
    sct_kwargs = {}
    for arg in sys.argv[1:]:
        if arg == '--no-auto':
            sct_kwargs['auto'] = False
        if arg == '--debug':
            sct_kwargs['debug'] = True
        if arg == '--no-suggestions':
            sct_kwargs['suggest'] = False
        if arg == '--no-google':
            sct_kwargs['google_sc'] = False
            if arg == '--no-spacy':
                sc_kwargs['_spacy'] = False
            
    sc = SpellChecker(**sc_kwargs)
    sc.spell_check_text(**sct_kwargs)
